from __future__ import annotations

"""Prompt for a database schema name and list all rows from its `inbridge` table."""

def check_bridge_restarts_raw(database: str, inbridgeid: int, limit: int = 100):
    """Voer direct het SQL-statement uit om bridge-restarts te tonen voor een specifieke inbridgeid."""
    conn = create_connection(database)
    if not conn:
        print(f"Kan niet verbinden met database {database}")
        return
    try:
        cur = conn.cursor(dictionary=True)
        sql = (
            "SELECT timestamp, "
            "inet_ntoa(conv(substr(comment,28,8), 16, 10)) as ip_address, "
            "conv(substr(comment,36,8),16,10) as currnt_time, "
            "timestamp - INTERVAL conv(substr(comment,36,8),16,10) SECOND as starttime "
            "FROM communicationlog "
            "WHERE comment like 'ab abab 55 5555 30 434f4e4e________________' "
            "AND inbridgeid=%s "
            "ORDER BY communicationlogid DESC LIMIT %s"
        )
        cur.execute(sql, (inbridgeid, limit))
        rows = cur.fetchall()
        if not rows:
            print(f"Geen restart-entries gevonden voor inbridgeid {inbridgeid}.")
            return
        print(f"\nBridge restarts voor inbridgeid {inbridgeid} (laatste {limit}):")
        print(f"{'timestamp':<20} {'ip_address':<15} {'currnt_time':<12} {'starttime':<20}")
        for r in rows:
            print(f"{r['timestamp']:<20} {r['ip_address']:<15} {r['currnt_time']:<12} {r['starttime']:<20}")
    except Exception as e:
        print(f"Fout bij uitvoeren van restart-check: {e}")
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
import os
import getpass
from dotenv import load_dotenv
import mysql.connector
import time
import argparse
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Font
import sys


# ANSI colors
GREEN = "\033[32m"
RESET = "\033[0m"


def _prepare_output(export_path: str | Path):
    """Prepare output directory and prefix from export_path.

    If export_path has a suffix (looks like a filename), its parent directory
    is created and the stem is returned as prefix. If export_path looks like
    a directory (no suffix), the directory is created and its name is used
    as prefix.
    Returns (parent_dir: Path, prefix: str).
    """
    p = Path(export_path)
    if p.suffix:
        parent = p.parent or Path('.')
        prefix = p.stem
    else:
        parent = p
        prefix = p.name or 'export'
    parent.mkdir(parents=True, exist_ok=True)
    return parent, prefix


def _apply_excel_sheet_styles(writer, sheet_name: str, df: pd.DataFrame):
    """Apply simple styles: add autofilter over header row, freeze header, and remove cell borders.

    `writer` is a pandas ExcelWriter using openpyxl engine.
    """
    try:
        wb = writer.book
        # writer.sheets populated after to_excel call
        ws = writer.sheets.get(sheet_name) or wb[sheet_name]
        # number of rows written (exclude header) — ensure at least header row
        nrows = max(0, len(df))
        ncols = max(1, len(df.columns))
        last_col = get_column_letter(ncols)
        # set autofilter across header row (A1:lastcol{nrows+1})
        ws.auto_filter.ref = f"A1:{last_col}{nrows+1}"
        # freeze header row
        ws.freeze_panes = ws['A2']
        # make header row bold
        try:
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
        except Exception:
            pass
        # remove borders from all cells in the used range (header + data)
        empty_border = Border()
        for row in ws.iter_rows(min_row=1, max_row=nrows+1, min_col=1, max_col=ncols):
            for cell in row:
                cell.border = empty_border
    except Exception:
        # non-fatal styling; ignore on failure
        pass


def _write_csv_with_fallback(path: Path, df: pd.DataFrame, sep: str = ';'):
    """Write `df` to `path`; on PermissionError retry with a timestamped filename.

    Returns the Path actually written or raises the original exception if both attempts fail.
    """
    try:
        df.to_csv(path, index=False, sep=sep, encoding='utf-8-sig')
        return path
    except Exception as e:
        # check for permission-like errors (PermissionError or errno 13)
        is_perm = isinstance(e, PermissionError) or getattr(e, 'errno', None) == 13
        if not is_perm:
            raise
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        alt = path.with_name(f"{path.stem}_{ts}{path.suffix}")
        try:
            df.to_csv(alt, index=False, sep=sep, encoding='utf-8-sig')
            print(f'Primary CSV {path} locked; wrote fallback CSV: {alt}')
            return alt
        except Exception:
            # re-raise original for caller to report
            raise


def _write_xlsx_with_fallback(path: Path, df: pd.DataFrame, sheet_name: str = 'Sheet1') -> Path:
    """Write `df` to `path` as an XLSX workbook, applying styles.

    On PermissionError (file locked), retry with a timestamped filename and
    return the Path actually written. Raises original exception if both fail.
    """
    try:
        with pd.ExcelWriter(path, engine='openpyxl') as ew:
            df.to_excel(ew, sheet_name=sheet_name, index=False)
            _apply_excel_sheet_styles(ew, sheet_name, df)
        return path
    except Exception as e:
        is_perm = isinstance(e, PermissionError) or getattr(e, 'errno', None) == 13
        if not is_perm:
            raise

        # If running interactively, ask user to close the file and retry.
        # If non-interactive (CI, background), automatically write a timestamped fallback.
        def _write_alt():
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            alt = path.with_name(f"{path.stem}_{ts}{path.suffix}")
            with pd.ExcelWriter(alt, engine='openpyxl') as ew:
                df.to_excel(ew, sheet_name=sheet_name, index=False)
                _apply_excel_sheet_styles(ew, sheet_name, df)
            print(f'Primary Excel {path} locked; wrote fallback Excel: {alt}')
            return alt

        if sys.stdin is None or not sys.stdin.isatty() or os.getenv('DBSCRIPT_NONINTERACTIVE') == '1':
            # non-interactive: fall back immediately
            try:
                return _write_alt()
            except Exception:
                raise

        # interactive prompt: let the user attempt to close and retry up to 3 times
        attempts = 0
        while attempts < 3:
            ans = input(f"File {path} appears locked. Close it and press Enter to retry, or type 's' to skip and write a timestamped fallback: ").strip().lower()
            if ans == 's':
                try:
                    return _write_alt()
                except Exception:
                    raise
            # otherwise, try to write again
            attempts += 1
            try:
                with pd.ExcelWriter(path, engine='openpyxl') as ew:
                    df.to_excel(ew, sheet_name=sheet_name, index=False)
                    _apply_excel_sheet_styles(ew, sheet_name, df)
                return path
            except Exception as inner_e:
                is_perm2 = isinstance(inner_e, PermissionError) or getattr(inner_e, 'errno', None) == 13
                if not is_perm2:
                    raise
                print(f"Still locked (attempt {attempts}).")

        # if we reach here, attempts exhausted — write fallback
        return _write_alt()


def _export_df_prompt(df: pd.DataFrame, export_path: Path | str):
    """Prompt to export `df` as XLSX, JSON, both, or none; optionally open after creation.

    `export_path` is treated as a prefix; files are written into the prepared directory.
    Returns dict of written file paths.
    """
    parent, prefix = _prepare_output(export_path)
    choice = input("Export format? [1] xlsx (default), [2] json, [3] both, [4] none: ").strip()
    if choice == '2':
        fmt = 'json'
    elif choice == '3':
        fmt = 'both'
    elif choice == '4':
        fmt = 'none'
    else:
        fmt = 'xlsx'

    written = {}
    if fmt in ('xlsx', 'both'):
        xlsxp = Path(parent) / f"{prefix}.xlsx"
        try:
            written_path = _write_xlsx_with_fallback(xlsxp, df, sheet_name='Logs')
            written['xlsx'] = written_path
            print(f'Wrote Excel workbook: {written_path}')
        except Exception as e:
            print(f'Failed to write Excel workbook {xlsxp}: {e}')

    if fmt in ('json', 'both'):
        jsonp = Path(parent) / f"{prefix}.json"
        try:
            df.to_json(jsonp, orient='records', indent=2, force_ascii=False)
            written['json'] = jsonp
            print(f'Wrote JSON file: {jsonp}')
        except PermissionError:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            alt = jsonp.with_name(f"{jsonp.stem}_{ts}{jsonp.suffix}")
            try:
                df.to_json(alt, orient='records', indent=2, force_ascii=False)
                written['json'] = alt
                print(f'Primary JSON {jsonp} locked; wrote fallback JSON: {alt}')
            except Exception as e:
                print(f'Failed to write JSON {jsonp}: {e}')
        except Exception as e:
            print(f'Failed to write JSON {jsonp}: {e}')

    if written:
        open_after = input('Open exported file(s) now? [y/N]: ').strip().lower() == 'y'
        if open_after:
            for p in written.values():
                try:
                    os.startfile(str(p))
                except Exception:
                    try:
                        os.startfile(str(p.parent))
                    except Exception:
                        pass
    else:
        print('No files written.')
    return written

def load_workspace_env():
    here = Path(__file__).resolve()
    candidates = [
        here.parent / ".env",
        here.parents[2] / ".env" if len(here.parents) >= 3 else None,
        here.parents[3] / "python" / "DBscript" / ".env" if len(here.parents) >= 4 else None,
        here.parents[3] / "DBscript" / ".env" if len(here.parents) >= 4 else None,
        Path.cwd() / ".env",
    ]

    for env_path in candidates:
        if env_path and env_path.exists():
            load_dotenv(env_path, override=True)
            return env_path

    load_dotenv()
    return None


ENV_PATH = load_workspace_env()

DB_HOST = os.getenv("DB_HOST")
DB_HOST2 = os.getenv("DB_HOST2")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

MAX_CONNECTION_ATTEMPTS = 3

if not DB_USER:
    DB_USER = input("DB user: ")
if not DB_PASSWORD:
    DB_PASSWORD = getpass.getpass("DB password: ")

def create_connection(database: str | None = None):
    hosts = [h for h in (DB_HOST, DB_HOST2) if h]
    if not hosts:
        hosts = ["localhost"]
    last_err = None
    for host in hosts:
        for attempt in range(MAX_CONNECTION_ATTEMPTS):
            try:
                conn = mysql.connector.connect(
                    host=host,
                    user=DB_USER,
                    password=DB_PASSWORD,
                    database=database,
                    connect_timeout=10,
                )
                if conn.is_connected():
                    print(f"{GREEN}OK \u2714{RESET} Connected to {host} (database={database})")
                    return conn
            except mysql.connector.Error as e:
                last_err = e
                # Keep individual attempts silent unless DB_DEBUG env var is set
                if os.getenv('DB_DEBUG'):
                    print(f"Connection attempt {attempt+1} to {host} failed: {e}")
            time.sleep(1)
    if last_err:
        print(f"Unable to connect to any host. Last error: {last_err}")
    else:
        print("Unable to connect to any host.")
    return None


def list_bridges_for_db(database: str):
    conn = create_connection(database)
    if not conn:
        print(f"Unable to connect to database {database}")
        return
    try:
        cur = conn.cursor()
        query = (
            "SELECT "
            "inbridgeid AS id, "
            "bridgetype, "
            "hostname AS macaddress, "
            "bridgestate AS status, "
            "errortext AS error, "
            "changetimestamp AS `last changed`, "
            "swversion AS `sw version`, "
            "polling, "
            "pollfailure, "
            "localip AS `IP-address local` "
            "FROM inbridge"
        )
        cur.execute(query)
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description] if cur.description else []
        if not rows:
            print("No rows returned from inbridge.")
            return
        # Nicely format table so values align under headers (toolkit style)
        def fmt_val(v):
            if v is None:
                return "NULL"
            return str(v)

        # compute column widths
        max_width = 60
        col_widths = []
        for i, col in enumerate(cols):
            col_len = len(str(col))
            max_len = col_len
            for r in rows:
                val_len = len(fmt_val(r[i]))
                if val_len > max_len:
                    max_len = val_len
            # cap width to avoid extremely wide columns
            if max_len > max_width:
                max_len = max_width
            col_widths.append(max_len)

        # Print toolkit-style header
        CYAN = "\033[36m"
        RESET = "\033[0m"
        print(CYAN + "Customers Bridges - {}".format(database) + RESET)
        # Header row
        header_cells = []
        for i, col in enumerate(cols):
            header_cells.append(str(col).ljust(col_widths[i]))
        print(" | ".join(header_cells))
        # Underline
        underline = []
        for w in col_widths:
            underline.append("-" * w)
        print("-|-".join(underline))

        # Rows
        for row in rows:
            cells = []
            for i, cell in enumerate(row):
                s = fmt_val(cell)
                # truncate if too long
                if len(s) > col_widths[i]:
                    s = s[: col_widths[i] - 3] + "..."
                cells.append(s.ljust(col_widths[i]))
            print(" | ".join(cells))
    except mysql.connector.Error as e:
        print(f"Query error: {e}")
    finally:
        cur.close()
        conn.close()


def analyze_all_bridges(database: str, gap_minutes: int = 15, restart_threshold: int = 3, limit: int = 100000, min_restart_days: int = 2, window_days: int = 4, restart_window_threshold: int = 20):
    """Analyze `communicationlog` for all bridges in `database`.

    Flags bridges with either more than `restart_window_threshold` restarts within
    the last `window_days` days, or with gaps longer than `gap_minutes` minutes.
    Returns a DataFrame of flagged bridges (may be empty).
    """
    conn = create_connection(database)
    if not conn:
        print(f"Unable to connect to database {database}")
        return pd.DataFrame()

    try:
        cur = conn.cursor(dictionary=True)
        q = (
            "SELECT inbridgeid, comment, timestamp FROM communicationlog "
            "WHERE inbridgeid IS NOT NULL ORDER BY inbridgeid, timestamp ASC LIMIT %s"
        )
        cur.execute(q, (limit,))
        rows = cur.fetchall()
        if not rows:
            print('No communicationlog rows found')
            return pd.DataFrame()

        # group rows by inbridgeid
        groups = {}
        for r in rows:
            bid = r.get('inbridgeid')
            if bid is None:
                continue
            groups.setdefault(bid, []).append(r)

        # helper classify
        def _classify_comment(comment: str) -> str:
            if not comment:
                return 'normal'
            lc = str(comment).lower()
            if '434f4e' in lc or '434f' in lc or 'conn' in lc:
                return 'restart'
            if 'abab' in lc:
                return 'ab'
            return 'normal'

        # Fetch bridge metadata (hostname/comment) to enrich output
        try:
            cur2 = conn.cursor(dictionary=True)
            cur2.execute('SELECT inbridgeid, hostname, comment FROM inbridge')
            meta_rows = cur2.fetchall()
            meta = {r['inbridgeid']: r for r in meta_rows}
        except Exception:
            meta = {}
        finally:
            try:
                cur2.close()
            except Exception:
                pass

        results = []
        cutoff = datetime.now() - timedelta(days=int(window_days))
        for bid, items in groups.items():
            # ensure sorted by timestamp
            try:
                items_sorted = sorted(items, key=lambda x: x.get('timestamp') or datetime.min)
            except Exception:
                items_sorted = items

            total = len(items_sorted)
            day_restart_counts = {}
            restart_count = 0
            restarts_in_window = 0
            ab_count = 0
            max_gap = 0.0
            gap_count = 0
            prev_ts = None

            for it in items_sorted:
                ts = it.get('timestamp')
                comment = it.get('comment')
                kind = _classify_comment(comment)
                if kind == 'restart':
                    restart_count += 1
                    try:
                        d = (ts.date() if hasattr(ts, 'date') else datetime.fromisoformat(str(ts)).date())
                        day_restart_counts[d] = day_restart_counts.get(d, 0) + 1
                        if ts is not None and ts >= cutoff:
                            restarts_in_window += 1
                    except Exception:
                        pass
                if kind == 'ab':
                    ab_count += 1

                if prev_ts is not None and ts is not None:
                    try:
                        delta_min = (ts - prev_ts).total_seconds() / 60.0
                        if delta_min > max_gap:
                            max_gap = delta_min
                        if delta_min > gap_minutes:
                            gap_count += 1
                    except Exception:
                        pass
                prev_ts = ts

            if day_restart_counts:
                max_restarts_in_day = max(day_restart_counts.values())
                date_of_max = next(d for d, c in day_restart_counts.items() if c == max_restarts_in_day)
                date_of_max_str = date_of_max.isoformat()
                days_over_threshold = sum(1 for c in day_restart_counts.values() if c >= restart_threshold)
            else:
                max_restarts_in_day = 0
                date_of_max_str = None
                days_over_threshold = 0

            row_meta = meta.get(bid, {})
            host = row_meta.get('hostname') if isinstance(row_meta, dict) else None
            comment_meta = row_meta.get('comment') if isinstance(row_meta, dict) else None
            results.append({
                'inbridgeid': bid,
                'host': host,
                'comment': comment_meta,
                'total': total,
                'restart': restart_count,
                'max_restarts_in_day': int(max_restarts_in_day),
                'date_max_restarts': date_of_max_str,
                'days_with_restarts_over_threshold': int(days_over_threshold),
                'restarts_in_window': int(restarts_in_window),
                'ab': ab_count,
                'gaps_over_threshold': gap_count,
                'max_gap_min': round(max_gap, 1),
            })

        df = pd.DataFrame(results)
        if df.empty:
            print(f"\nNo communicationlog rows for database {database}")
            return pd.DataFrame()

        # Flag only those with restarts in window above threshold OR gaps above gap_minutes
        if 'restarts_in_window' not in df.columns:
            df['restarts_in_window'] = df.get('restart', 0)
        mask = ((df['restarts_in_window'] > int(restart_window_threshold)) | (df['max_gap_min'] > gap_minutes))
        flagged_df = df[mask].copy()

        print(f"\nBridge health summary (db={database}) — gap threshold {gap_minutes} min, window_days={window_days}, restart_window_threshold={restart_window_threshold}")
        if flagged_df.empty:
            # return empty dataframe; caller prints approval when desired
            return flagged_df

        disp_cols = ['inbridgeid', 'host', 'total', 'restart', 'restarts_in_window', 'max_restarts_in_day', 'date_max_restarts', 'ab', 'gaps_over_threshold', 'max_gap_min']
        # Poll fails percentage (if possible)
        try:
            cur2 = conn.cursor(dictionary=True)
            cur2.execute('SELECT inbridgeid, polling, pollfailure FROM inbridge')
            poll_rows = cur2.fetchall()
            poll_map = {r['inbridgeid']: r for r in poll_rows}
        except Exception:
            poll_map = {}
        finally:
            try:
                cur2.close()
            except Exception:
                pass
        poll_fail_perc = []
        for ix, row in flagged_df.iterrows():
            bid = row['inbridgeid']
            poll = poll_map.get(bid, {}).get('polling', 0)
            pollfail = poll_map.get(bid, {}).get('pollfailure', 0)
            try:
                poll = int(poll)
                pollfail = int(pollfail)
                perc = 100.0 * pollfail / poll if poll else 0.0
            except Exception:
                perc = 0.0
            poll_fail_perc.append(round(perc, 1))
        flagged_df['pollfail_percent'] = poll_fail_perc
        disp_cols.append('pollfail_percent')
        # Filter for pollfail > 15% if any
        pollfail_flagged = flagged_df[flagged_df['pollfail_percent'] > 15.0]
        if not pollfail_flagged.empty:
            print('\nBridges with >15% poll fails:')
            print(pollfail_flagged.to_string(index=False))
        for c in disp_cols:
            if c not in flagged_df.columns:
                flagged_df[c] = ''
        flagged_df = flagged_df[disp_cols]
        print(flagged_df.sort_values(['restarts_in_window', 'max_gap_min'], ascending=False).to_string(index=False))
        return flagged_df

    except mysql.connector.Error as e:
        print(f"Query error: {e}")
        return pd.DataFrame()
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def analyze_all_databases(gap_minutes: int = 15, restart_threshold: int = 3, limit: int = 100000, include_system: bool = False, export_path: str | None = None, min_restart_days: int = 2, window_days: int = 4, restart_window_threshold: int = 20):
    """Scan all databases on configured hosts and run `analyze_all_bridges` for those
    that contain an `inbridge` table.

    Skips system schemas by default. Connects using the same `create_connection` helper
    (no default database) so `.env` hosts and credentials are used.
    """
    conn = create_connection(None)
    if not conn:
        print('Unable to connect to any host to list databases')
        return
    try:
        cur = conn.cursor()
        cur.execute('SHOW DATABASES')
        dbs = [r[0] for r in cur.fetchall()]
    except Exception as e:
        print(f'Failed to list databases: {e}')
        return
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    skip = {'mysql', 'information_schema', 'performance_schema', 'sys'}
    all_flagged = {}
    for db in sorted(dbs):
        if not include_system and db in skip:
            continue
        # Quick check whether this schema contains an inbridge table
        try:
            c = create_connection(db)
            if not c:
                print(f"Skipping {db}: cannot connect")
                continue
            try:
                cur2 = c.cursor()
                cur2.execute("SELECT 1 FROM inbridge LIMIT 1")
                cur2.fetchall()
                has_inbridge = True
            except Exception:
                has_inbridge = False
            finally:
                try:
                    cur2.close()
                except Exception:
                    pass
                try:
                    c.close()
                except Exception:
                    pass
        except Exception:
            has_inbridge = False

        if not has_inbridge:
            # skip schemas without inbridge table
            continue

        print('\n' + '=' * 60)
        print(f"Analyzing database: {db}")
        print('=' * 60)
        flagged_df = analyze_all_bridges(db, gap_minutes=gap_minutes, restart_threshold=restart_threshold, limit=limit, min_restart_days=min_restart_days, window_days=window_days, restart_window_threshold=restart_window_threshold)
        if flagged_df is not None and not flagged_df.empty:
            all_flagged[db] = flagged_df
        else:
            # no issues found — print green approval and skip exporting
            print(f"{GREEN}OK \u2714{RESET} — {db} has no restarts (> {restart_window_threshold} in {window_days}d) or gaps (> {gap_minutes} min)")

    # After scanning all DBs, optionally summarize and offer export helpers
    if not all_flagged:
        print('\nNo problematic bridges found across scanned databases.')
        return all_flagged

    # Combine for overview (one table, one sheet)
    combined = pd.concat([df.assign(database=db) for db, df in all_flagged.items()], ignore_index=True)
    # Reorder columns
    cols = ['database'] + [c for c in combined.columns if c != 'database']
    combined = combined[cols]
    print('\nCombined flagged bridges across databases:')
    print(combined.to_string(index=False))

    # Export combined results if requested (one sheet, one table)
    if export_path:
        try:
            parent, prefix = _prepare_output(export_path)
            csvp = parent / f"{prefix}.csv"
            # interactive export: prefer xlsx/json instead of CSV
            try:
                _export_df_prompt(combined, csvp)
            except Exception as e:
                print(f'Export failed: {e}')
        except Exception as e:
            print(f'Failed to prepare CSV path: {e}')

        xlsxp = parent / f"{prefix}.xlsx"
        try:
            written_path = _write_xlsx_with_fallback(xlsxp, combined, sheet_name='FlaggedBridges')
            print(f'Wrote Excel workbook: {written_path}')
        except Exception as e:
            print(f'Excel export failed (openpyxl may be missing or file locked): {e}')

    return all_flagged


def analyze_poll_failures_db(database: str, threshold: int = 10, days: int = 1):
    """Return DataFrame of bridges in `database` where pollfailure > threshold
    AND where `bridgestate` is OPEN or `changetimestamp` is within `days` days.
    """
    conn = create_connection(database)
    if not conn:
        print(f"Unable to connect to database {database}")
        return pd.DataFrame()
    try:
        cur = conn.cursor(dictionary=True)
        q = (
            "SELECT inbridgeid, hostname, polling, pollfailure, bridgestate, changetimestamp, comment "
            "FROM inbridge"
        )
        cur.execute(q)
        rows = cur.fetchall()
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        # ensure numeric pollfailure
        if 'pollfailure' in df.columns:
            df['pollfailure'] = pd.to_numeric(df['pollfailure'], errors='coerce').fillna(0).astype(int)
        else:
            df['pollfailure'] = 0

        # normalize changetimestamp and bridgestate
        if 'changetimestamp' in df.columns:
            df['changetimestamp'] = pd.to_datetime(df['changetimestamp'], errors='coerce')
        else:
            df['changetimestamp'] = pd.NaT
        if 'bridgestate' in df.columns:
            df['bridgestate_norm'] = df['bridgestate'].astype(str).str.strip().str.upper()
        else:
            df['bridgestate_norm'] = ''

        cutoff = pd.Timestamp(datetime.now() - timedelta(days=int(days)))

        # base mask: pollfailure > threshold
        base_mask = df['pollfailure'] > int(threshold)
        # additional recent/open mask
        recent_open_mask = (df['bridgestate_norm'] == 'OPEN') | (df['changetimestamp'] >= cutoff)

        # require both: pollfailure exceeded AND (open or recent)
        flagged = df[base_mask & recent_open_mask].copy()

        # compute percentage of poll failures and sort high->low
        try:
            def _safe_pct_row(r):
                try:
                    poll = int(r.get('polling', 0) or 0)
                    fail = int(r.get('pollfailure', 0) or 0)
                    return round(100.0 * fail / poll, 1) if poll else 0.0
                except Exception:
                    return 0.0

            flagged['pollfail_percent'] = flagged.apply(_safe_pct_row, axis=1)
        except Exception:
            flagged['pollfail_percent'] = 0.0

        if not flagged.empty:
            flagged = flagged.sort_values('pollfail_percent', ascending=False)
            print(f"\nPoll-fail summary (db={database}) threshold={threshold}, recent_days={days}")
            cols = ['inbridgeid', 'hostname', 'polling', 'pollfailure', 'pollfail_percent', 'bridgestate', 'changetimestamp']
            available = [c for c in cols if c in flagged.columns]
            print(flagged[available].to_string(index=False))
        return flagged
    except mysql.connector.Error as e:
        print(f"Query error: {e}")
        return pd.DataFrame()
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def analyze_poll_failures_all(threshold: int = 10, days: int = 1, include_system: bool = False, export_path: str | None = None):
    """Scan all databases and collect bridges with pollfailure > threshold.

    Returns mapping database->DataFrame for databases with flagged rows. Optionally exports combined CSV/XLSX when export_path given.
    """
    conn = create_connection(None)
    if not conn:
        print('Unable to connect to any host to list databases')
        return {}
    try:
        cur = conn.cursor()
        cur.execute('SHOW DATABASES')
        dbs = [r[0] for r in cur.fetchall()]
    except Exception as e:
        print(f'Failed to list databases: {e}')
        return {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    skip = {'mysql', 'information_schema', 'performance_schema', 'sys'}
    results = {}
    for db in sorted(dbs):
        if not include_system and db in skip:
            continue
        # quick check for inbridge existence
        try:
            c = create_connection(db)
            if not c:
                continue
            try:
                cur2 = c.cursor()
                cur2.execute('SELECT 1 FROM inbridge LIMIT 1')
                has = True
            except Exception:
                has = False
            finally:
                try:
                    cur2.close()
                except Exception:
                    pass
                try:
                    c.close()
                except Exception:
                    pass
        except Exception:
            has = False
        if not has:
            continue
        flagged = analyze_poll_failures_db(db, threshold=threshold, days=days)
        if flagged is not None and not flagged.empty:
            results[db] = flagged

    if not results:
        print('\nNo poll-failures above threshold found across scanned databases.')
        return results

    combined = pd.concat([df.assign(database=db) for db, df in results.items()], ignore_index=True)
    # ensure pollfail percent exists and sort by it (high->low)
    try:
        if 'pollfail_percent' not in combined.columns:
            def _safe_pct_row(r):
                try:
                    poll = int(r.get('polling', 0) or 0)
                    fail = int(r.get('pollfailure', 0) or 0)
                    return round(100.0 * fail / poll, 1) if poll else 0.0
                except Exception:
                    return 0.0
            combined['pollfail_percent'] = combined.apply(_safe_pct_row, axis=1)
    except Exception:
        combined['pollfail_percent'] = 0.0

    combined = combined.sort_values('pollfail_percent', ascending=False)
    print('\nCombined poll-fail results:')
    cols_print = [c for c in ['database', 'inbridgeid', 'hostname', 'polling', 'pollfailure', 'pollfail_percent'] if c in combined.columns]
    print(combined[cols_print].to_string(index=False))

    # Always export to ICY-Logs
    export_dir = Path(r'C:/Users/h.nijdam/Documents/ICY-Logs')
    export_dir.mkdir(parents=True, exist_ok=True)
    csvp = export_dir / 'pollfail_combined.csv'
    xlsxp = export_dir / 'pollfail_combined.xlsx'
    try:
        # interactive export: prefer xlsx/json instead of CSV
        try:
            _export_df_prompt(combined, csvp)
        except Exception as e:
            print(f'Export failed: {e}')
    except Exception as e:
        print(f'Failed to prepare or export combined pollfail files: {e}')
    try:
        written_path = _write_xlsx_with_fallback(xlsxp, combined, sheet_name='PollFails')
        print(f'Wrote combined pollfail workbook: {written_path}')
    except Exception as e:
        print(f'Excel export failed for pollfails (openpyxl may be missing or file locked): {e}')

    return results


def analyze_open_recent_all(days: int = 1, include_system: bool = False, export_path: str | None = None):
    """Scan all databases and collect bridges where bridgestate is OPEN or last changed within `days` days.

    Writes per-DB sheets into a combined XLSX when `export_path` provided.
    Returns mapping database->DataFrame for databases with flagged rows.
    """
    conn = create_connection(None)
    if not conn:
        print('Unable to connect to any host to list databases')
        return {}
    try:
        cur = conn.cursor()
        cur.execute('SHOW DATABASES')
        dbs = [r[0] for r in cur.fetchall()]
    except Exception as e:
        print(f'Failed to list databases: {e}')
        return {}
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    skip = {'mysql', 'information_schema', 'performance_schema', 'sys'}
    results = {}
    cutoff = datetime.now() - timedelta(days=int(days))
    for db in sorted(dbs):
        if not include_system and db in skip:
            continue
        # quick check for inbridge existence
        try:
            c = create_connection(db)
            if not c:
                continue
            try:
                cur2 = c.cursor(dictionary=True)
                cur2.execute('SELECT inbridgeid, hostname, bridgestate, changetimestamp, polling, pollfailure, comment FROM inbridge')
                rows = cur2.fetchall()
            except Exception:
                rows = []
            finally:
                try:
                    cur2.close()
                except Exception:
                    pass
                try:
                    c.close()
                except Exception:
                    pass
        except Exception:
            rows = []

        if not rows:
            continue
        df = pd.DataFrame(rows)
        if df.empty:
            continue
        # normalize changetimestamp
        if 'changetimestamp' in df.columns:
            df['changetimestamp'] = pd.to_datetime(df['changetimestamp'], errors='coerce')
        else:
            df['changetimestamp'] = pd.NaT
        # bridgestate normalize
        if 'bridgestate' in df.columns:
            df['bridgestate_norm'] = df['bridgestate'].astype(str).str.strip().str.upper()
        else:
            df['bridgestate_norm'] = ''

        mask = (df['bridgestate_norm'] == 'OPEN') | (df['changetimestamp'] >= pd.Timestamp(cutoff))
        flagged = df[mask].copy()
        if not flagged.empty:
            results[db] = flagged
            print(f"\nOpen/recent bridges (db={db}) — found {len(flagged)} rows")
            print(flagged[['inbridgeid', 'hostname', 'bridgestate', 'changetimestamp']].to_string(index=False))
            if export_path:
                pass

    if not results:
        print('\nNo open or recently changed bridges found across scanned databases.')
        return results

    combined = pd.concat([df.assign(database=db) for db, df in results.items()], ignore_index=True)
    print('\nCombined open/recent results:')
    print(combined[['database', 'inbridgeid', 'hostname', 'bridgestate', 'changetimestamp']].to_string(index=False))

    if export_path:
        try:
            csvp = Path(export_path).with_suffix('.openrecent')
            try:
                # interactive export prompt (xlsx/json/both)
                _export_df_prompt(combined, csvp)
            except Exception as e:
                print(f'Export failed: {e}')
        except Exception as e:
            print(f'Failed to prepare openrecent export path: {e}')

    return results




def main_menu():
    options = [
        "List databases",
        "Enumerate all bridges (all DBs)",
        "Select a database and manage bridges",
        "Exit"
    ]
    # This helper script shouldn't attempt to use the interactive selector
    # from `db_menu.py`. Keep a simple prompt-based entrypoint below.


def main():
    # legacy interactive entrypoint preserved for backward compatibility
    db = input("Database name (or blank to exit): ").strip()
    if not db:
        print("No database provided. Exiting.")
        return
    action = input("Enter 'a' to analyze this DB, 'poll' to check poll-failures, 'all' to analyze all databases, 'pollall' to check poll-failures across all DBs, or press Enter to list bridges: ").strip().lower()
    if action == 'a' or action == 'analyze':
        try:
            gm = input("Gap threshold minutes (default 15): ").strip()
            gap = int(gm) if gm else 15
        except Exception:
            gap = 15
        try:
            rt = input("Restart alert threshold (count, default 3): ").strip()
            rtv = int(rt) if rt else 3
        except Exception:
            rtv = 3
        try:
            lim = input("Limit rows from communicationlog to fetch (default 100000): ").strip()
            limv = int(lim) if lim else 100000
        except Exception:
            limv = 100000
        try:
            mrd = input("Min distinct days with restarts to flag (default 2): ").strip()
            mrdv = int(mrd) if mrd else 2
        except Exception:
            mrdv = 2
        df = analyze_all_bridges(db, gap_minutes=gap, restart_threshold=rtv, limit=limv, min_restart_days=mrdv, window_days=4, restart_window_threshold=20)
        if df is not None and df.empty:
            print(f"{GREEN}GOED! \u2714{RESET} — {db} has no restarts (>20 in 4d) or gaps (> {gap} min)")
    elif action == 'poll':
        try:
            th = input("Poll-failure threshold (default 10): ").strip()
            thv = int(th) if th else 10
        except Exception:
            thv = 10
        exp = input("Export path prefix (optional, writes per-db and combined files): ").strip() or None
        analyze_poll_failures_db(db, threshold=thv)
    elif action == 'all':
        try:
            gm = input("Gap threshold minutes (default 15): ").strip()
            gap_all = int(gm) if gm else 15
        except Exception:
            gap_all = 15
        try:
            rt = input("Restart alert threshold (count, default 3): ").strip()
            rtv_all = int(rt) if rt else 3
        except Exception:
            rtv_all = 3
        try:
            mrd_all = input("Min distinct days with restarts to flag (default 2): ").strip()
            mrd_all_v = int(mrd_all) if mrd_all else 2
        except Exception:
            mrd_all_v = 2
        exp_choice = input("Export combined XLSX path prefix (optional, e.g. ./bridge_health_report): ").strip() or None
        analyze_all_databases(gap_minutes=gap_all, restart_threshold=rtv_all, limit=100000, export_path=exp_choice, min_restart_days=mrd_all_v, window_days=4, restart_window_threshold=20)
    elif action == 'pollall':
        try:
            th = input("Poll-failure threshold for all DBs (default 10): ").strip()
            th_all = int(th) if th else 10
        except Exception:
            th_all = 10
        exp_all = input("Export path prefix for pollfail results (optional): ").strip() or None
        analyze_poll_failures_all(threshold=th_all, include_system=False, export_path=exp_all)
    else:
        list_bridges_for_db(db)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Bridge health tooling: list/analyze bridges and poll-failures')
    parser.add_argument('--db', help='Database/schema name to operate on')
    parser.add_argument('--action', choices=['list', 'analyze', 'poll', 'all', 'pollall', 'openrecent'], help="Action: list, analyze, poll, all, pollall, openrecent")
    parser.add_argument('--gap-minutes', type=int, default=15, help='Gap threshold in minutes (default 15)')
    parser.add_argument('--restart-threshold', type=int, default=3, help='Restart alert threshold in a single day (default 3)')
    parser.add_argument('--limit', type=int, default=100000, help='Row limit when scanning communicationlog (default 100000)')
    parser.add_argument('--min-restart-days', type=int, default=2, help='Min distinct calendar days with restarts to flag (default 2)')
    parser.add_argument('--poll-threshold', type=int, default=10, help='Poll-failure count threshold (default 10)')
    parser.add_argument('--recent-days', type=int, default=1, help='Days to treat a change as recent for poll/open filtering (default 1)')
    parser.add_argument('--window-days', type=int, default=4, help='Window in days to count restarts (default 4)')
    parser.add_argument('--restart-window-threshold', type=int, default=20, help='Restart count threshold within window-days to flag (default 20)')
    parser.add_argument('--export', help='Export path prefix for writing CSV/XLSX outputs (optional)')
    args = parser.parse_args()

    # If no CLI args provided, fall back to interactive
    if not any([args.db, args.action, args.export]):
        main()
    else:
        act = (args.action or 'list')
        if act == 'list':
            if not args.db:
                parser.error("--db is required for 'list' action")
            list_bridges_for_db(args.db)
        elif act == 'analyze':
            if args.db:
                df = analyze_all_bridges(args.db, gap_minutes=args.gap_minutes, restart_threshold=args.restart_threshold, limit=args.limit, min_restart_days=args.min_restart_days, window_days=args.window_days, restart_window_threshold=args.restart_window_threshold)
                if df is not None and df.empty:
                    print(f"{GREEN}OK! \u2714{RESET} — {args.db} has no restarts (>{args.restart_window_threshold} in {args.window_days}d) or gaps (> {args.gap_minutes} min)")
            else:
                analyze_all_databases(gap_minutes=args.gap_minutes, restart_threshold=args.restart_threshold, limit=args.limit, export_path=args.export, min_restart_days=args.min_restart_days, window_days=args.window_days, restart_window_threshold=args.restart_window_threshold)
        elif act == 'poll':
            if args.db:
                analyze_poll_failures_db(args.db, threshold=args.poll_threshold, days=args.recent_days)
            else:
                parser.error("--db is required for 'poll' action")
        elif act == 'pollall':
            analyze_poll_failures_all(threshold=args.poll_threshold, days=args.recent_days, include_system=False, export_path=args.export)
        elif act == 'all':
            analyze_all_databases(gap_minutes=args.gap_minutes, restart_threshold=args.restart_threshold, limit=args.limit, export_path=args.export, min_restart_days=args.min_restart_days, window_days=args.window_days, restart_window_threshold=args.restart_window_threshold)
        elif act == 'openrecent':
            # days parameter implicit from min_restart_days CLI for convenience
            days = args.min_restart_days if args.min_restart_days else 1
            analyze_open_recent_all(days=days, include_system=False, export_path=args.export)
